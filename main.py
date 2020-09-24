from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
ws.title = "Worksheet 1"
ws.sheet_properties.tabColor = "1072BA"
print(wb.sheetnames)

#ws3 = wb["New Title"]

#copying a worksheet
#>>> source = wb.active
#>>> target = wb.copy_worksheet(source)

#fetch a cell
c = ws['A4']
#set a cell
ws['A4'] = 4
#There is also the Worksheet.cell() method.
#This provides access to cells using row and column notation:
#>>> d = ws.cell(row=4, column=2, value=10)

#ranges of cells can be accessed using slicing:
#>>> cell_range = ws['A1':'C2']

'''Ranges of rows or columns can be obtained similarly:

>>> colC = ws['C']
>>> col_range = ws['C:D']
>>> row10 = ws[10]
>>> row_range = ws[5:10]
'''
wb.save('balances.xlsx')
